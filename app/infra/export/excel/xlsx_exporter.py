
from __future__ import annotations

import time
from typing import Any, Iterable, Mapping, Sequence, Tuple

from app.core.ports.table_exporter_port import ExportMeta, ExportResult, TableExporterPort


class XlsxTableExporter(TableExporterPort):
    def export(
        self,
        rows: Iterable[Mapping[str, Any]],
        columns: Sequence[Tuple[str, str]],
        meta: ExportMeta,
        destination_path: str,
    ) -> ExportResult:
        t0 = time.time()
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
        except Exception as e:
            raise RuntimeError("Dependência 'openpyxl' não disponível para exportar XLSX.") from e

        wb = Workbook()
        ws = wb.active
        ws.title = "Dados"

        ws["A1"] = meta.title
        ws["A2"] = f"Gerado em: {meta.generated_at_iso}"
        ws["A3"] = f"Total de itens: {meta.total_rows}"
        if meta.query_summary:
            ws["A4"] = f"Consulta: {meta.query_summary}"

        ws["A1"].font = Font(size=14, bold=True)
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
        for r in ("A2", "A3", "A4"):
            ws[r].alignment = Alignment(horizontal="left", vertical="center")

        header_row = 6
        for c, (_, title) in enumerate(columns, start=1):
            cell = ws.cell(row=header_row, column=c, value=title)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.fill = PatternFill("solid", fgColor="DDDDDD")

        ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

        row_idx = header_row + 1
        exported = 0
        max_lens = [len(t) for _, t in columns]

        for r in rows:
            for c, (key, _) in enumerate(columns, start=1):
                v = r.get(key, "")
                if v is None:
                    v = ""
                ws.cell(row=row_idx, column=c, value=v)
                s = str(v)
                if len(s) > max_lens[c-1]:
                    max_lens[c-1] = len(s)
            row_idx += 1
            exported += 1

        for i, ml in enumerate(max_lens, start=1):
            w = max(10, min(60, ml + 2))
            ws.column_dimensions[_col_letter(i)].width = w

        wb.save(destination_path)
        return ExportResult(path=destination_path, rows_exported=exported, duration_ms=int((time.time()-t0)*1000))


def _col_letter(n: int) -> str:
    s = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s
